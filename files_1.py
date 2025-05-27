import openpyxl
import datetime
import os
import sqlite3
from flask import flash


def update_excel_template(user):


    # Формирование имени файла
    file_name = f'v_{datetime.datetime.today().strftime("%Y-%m-%d_%H-%M-%S")}.xlsx'

    # Получение текущей директории файла и добавление подкаталога `tmp`
    current_dir = os.path.abspath(os.path.dirname(__file__))
    tmp_dir = os.path.join(current_dir, 'tmp')

    # Убедитесь, что директория `tmp` существует
    os.makedirs(tmp_dir, exist_ok=True)

    # Полный путь к файлу в подкаталоге `tmp`
    file_path = os.path.join(tmp_dir, file_name)
    kvartal = 1
    try:
        # Подключение к базе данных
        db_lp = sqlite3.connect('date_source.db')
        cursor_db = db_lp.cursor()

        text = '''
                           SELECT
                        spisok_in_studio.napravlenie,
                        teacher.FIO AS teacher_name,
                        SUM(CASE WHEN events_table.level = 'Центровский' AND data_table.result = 'Сертификат участника' AND  strftime('%m', events_table.result_date) = '01' THEN 1 ELSE 0 END) AS c1,
                        SUM(CASE WHEN events_table.level = 'Центровский' AND data_table.result != 'Сертификат участника' AND  strftime('%m', events_table.result_date) = '01'  THEN 1 ELSE 0 END) AS c2,
                         SUM(CASE WHEN events_table.level = 'Центровский' AND data_table.result = 'Сертификат участника' AND  strftime('%m', events_table.result_date) = '02' THEN 1 ELSE 0 END) AS c12,
                        SUM(CASE WHEN events_table.level = 'Центровский' AND data_table.result != 'Сертификат участника' AND  strftime('%m', events_table.result_date) = '02'  THEN 1 ELSE 0 END) AS c22,
                         SUM(CASE WHEN events_table.level = 'Центровский' AND data_table.result = 'Сертификат участника' AND  strftime('%m', events_table.result_date) = '03' THEN 1 ELSE 0 END) AS c13,
                        SUM(CASE WHEN events_table.level = 'Центровский' AND data_table.result != 'Сертификат участника' AND  strftime('%m', events_table.result_date) = '03'  THEN 1 ELSE 0 END) AS c23,
                        SUM(CASE WHEN events_table.level = 'Городской' AND data_table.result = 'Сертификат участника' AND  strftime('%m', events_table.result_date) = '01'  THEN 1 ELSE 0 END) AS c3,
                        SUM(CASE WHEN events_table.level = 'Городской' AND data_table.result != 'Сертификат участника' AND  strftime('%m', events_table.result_date) = '01'  THEN 1 ELSE 0 END) AS c4,
                        SUM(CASE WHEN events_table.level = 'Городской' AND data_table.result = 'Сертификат участника' AND  strftime('%m', events_table.result_date) = '02'  THEN 1 ELSE 0 END) AS c32,
                        SUM(CASE WHEN events_table.level = 'Городской' AND data_table.result != 'Сертификат участника' AND  strftime('%m', events_table.result_date) = '02'  THEN 1 ELSE 0 END) AS c42,
                        SUM(CASE WHEN events_table.level = 'Городской' AND data_table.result = 'Сертификат участника' AND  strftime('%m', events_table.result_date) = '03'  THEN 1 ELSE 0 END) AS c33,
                        SUM(CASE WHEN events_table.level = 'Городской' AND data_table.result != 'Сертификат участника' AND  strftime('%m', events_table.result_date) = '03'  THEN 1 ELSE 0 END) AS c43,
                        SUM(CASE WHEN events_table.level = 'Районный' AND data_table.result = 'Сертификат участника' AND  strftime('%m', events_table.result_date) = '01'  THEN 1 ELSE 0 END) AS c5,
                        SUM(CASE WHEN events_table.level = 'Районный' AND data_table.result != 'Сертификат участника' AND  strftime('%m', events_table.result_date) = '01'  THEN 1 ELSE 0 END) AS c6,
                        SUM(CASE WHEN events_table.level = 'Районный' AND data_table.result = 'Сертификат участника' AND  strftime('%m', events_table.result_date) = '02'  THEN 1 ELSE 0 END) AS c52,
                        SUM(CASE WHEN events_table.level = 'Районный' AND data_table.result != 'Сертификат участника' AND  strftime('%m', events_table.result_date) = '02'  THEN 1 ELSE 0 END) AS c62,
                        SUM(CASE WHEN events_table.level = 'Районный' AND data_table.result = 'Сертификат участника' AND  strftime('%m', events_table.result_date) = '03'  THEN 1 ELSE 0 END) AS c53,
                        SUM(CASE WHEN events_table.level = 'Районный' AND data_table.result != 'Сертификат участника' AND  strftime('%m', events_table.result_date) = '03'  THEN 1 ELSE 0 END) AS c63,
                        SUM(CASE WHEN events_table.level = 'Республиканский' AND data_table.result = 'Сертификат участника' AND  strftime('%m', events_table.result_date) = '01'  THEN 1 ELSE 0 END) AS c7,
                        SUM(CASE WHEN events_table.level = 'Республиканский' AND data_table.result != 'Сертификат участника' AND  strftime('%m', events_table.result_date) = '01'  THEN 1 ELSE 0 END) AS c8,
                        SUM(CASE WHEN events_table.level = 'Республиканский' AND data_table.result = 'Сертификат участника' AND  strftime('%m', events_table.result_date) = '02'  THEN 1 ELSE 0 END) AS c72,
                        SUM(CASE WHEN events_table.level = 'Республиканский' AND data_table.result != 'Сертификат участника' AND  strftime('%m', events_table.result_date) = '02'  THEN 1 ELSE 0 END) AS c82,
                        SUM(CASE WHEN events_table.level = 'Республиканский' AND data_table.result = 'Сертификат участника' AND  strftime('%m', events_table.result_date) = '03'  THEN 1 ELSE 0 END) AS c73,
                        SUM(CASE WHEN events_table.level = 'Республиканский' AND data_table.result != 'Сертификат участника' AND  strftime('%m', events_table.result_date) = '03'  THEN 1 ELSE 0 END) AS c83,
                        SUM(CASE WHEN events_table.level = 'Региональный' AND data_table.result = 'Сертификат участника' AND  strftime('%m', events_table.result_date) = '01'  THEN 1 ELSE 0 END) AS c9,
                        SUM(CASE WHEN events_table.level = 'Региональный' AND data_table.result != 'Сертификат участника' AND  strftime('%m', events_table.result_date) = '01'  THEN 1 ELSE 0 END) AS c10,
                         SUM(CASE WHEN events_table.level = 'Региональный' AND data_table.result = 'Сертификат участника' AND  strftime('%m', events_table.result_date) = '02'  THEN 1 ELSE 0 END) AS c92,
                        SUM(CASE WHEN events_table.level = 'Региональный' AND data_table.result != 'Сертификат участника' AND  strftime('%m', events_table.result_date) = '02'  THEN 1 ELSE 0 END) AS c102,
                         SUM(CASE WHEN events_table.level = 'Региональный' AND data_table.result = 'Сертификат участника' AND  strftime('%m', events_table.result_date) = '03'  THEN 1 ELSE 0 END) AS c93,
                        SUM(CASE WHEN events_table.level = 'Региональный' AND data_table.result != 'Сертификат участника' AND  strftime('%m', events_table.result_date) = '03'  THEN 1 ELSE 0 END) AS c103,
                        SUM(CASE WHEN events_table.level = 'Межрегиональный' AND data_table.result = 'Сертификат участника' AND  strftime('%m', events_table.result_date) = '01'  THEN 1 ELSE 0 END) AS c11,
                        SUM(CASE WHEN events_table.level = 'Межрегиональный' AND data_table.result != 'Сертификат участника' AND  strftime('%m', events_table.result_date) = '01'  THEN 1 ELSE 0 END) AS c12,
                         SUM(CASE WHEN events_table.level = 'Межрегиональный' AND data_table.result = 'Сертификат участника' AND  strftime('%m', events_table.result_date) = '02'  THEN 1 ELSE 0 END) AS c112,
                        SUM(CASE WHEN events_table.level = 'Межрегиональный' AND data_table.result != 'Сертификат участника' AND  strftime('%m', events_table.result_date) = '02'  THEN 1 ELSE 0 END) AS c122,
                         SUM(CASE WHEN events_table.level = 'Межрегиональный' AND data_table.result = 'Сертификат участника' AND  strftime('%m', events_table.result_date) = '03'  THEN 1 ELSE 0 END) AS c113,
                        SUM(CASE WHEN events_table.level = 'Межрегиональный' AND data_table.result != 'Сертификат участника' AND  strftime('%m', events_table.result_date) = '03'  THEN 1 ELSE 0 END) AS c123,
                        SUM(CASE WHEN events_table.level = 'Всероссийский' AND data_table.result = 'Сертификат участника' AND  strftime('%m', events_table.result_date) = '01'  THEN 1 ELSE 0 END) AS c13,
                        SUM(CASE WHEN events_table.level = 'Всероссийский' AND data_table.result != 'Сертификат участника' AND  strftime('%m', events_table.result_date) = '01'  THEN 1 ELSE 0 END) AS c14,
                        SUM(CASE WHEN events_table.level = 'Всероссийский' AND data_table.result = 'Сертификат участника' AND  strftime('%m', events_table.result_date) = '02'  THEN 1 ELSE 0 END) AS c132,
                        SUM(CASE WHEN events_table.level = 'Всероссийский' AND data_table.result != 'Сертификат участника' AND  strftime('%m', events_table.result_date) = '02'  THEN 1 ELSE 0 END) AS c142,
                        SUM(CASE WHEN events_table.level = 'Всероссийский' AND data_table.result = 'Сертификат участника' AND  strftime('%m', events_table.result_date) = '03'  THEN 1 ELSE 0 END) AS c133,
                        SUM(CASE WHEN events_table.level = 'Всероссийский' AND data_table.result != 'Сертификат участника' AND  strftime('%m', events_table.result_date) = '03'  THEN 1 ELSE 0 END) AS c143,
                        SUM(CASE WHEN events_table.level = 'Международный' AND data_table.result = 'Сертификат участника' AND  strftime('%m', events_table.result_date) = '01'  THEN 1 ELSE 0 END) AS c15,
                        SUM(CASE WHEN events_table.level = 'Международный' AND data_table.result != 'Сертификат участника' AND  strftime('%m', events_table.result_date) = '01'  THEN 1 ELSE 0 END) AS c16,
                        SUM(CASE WHEN events_table.level = 'Международный' AND data_table.result = 'Сертификат участника' AND  strftime('%m', events_table.result_date) = '02'  THEN 1 ELSE 0 END) AS c152,
                        SUM(CASE WHEN events_table.level = 'Международный' AND data_table.result != 'Сертификат участника' AND  strftime('%m', events_table.result_date) = '02'  THEN 1 ELSE 0 END) AS c162,
                        SUM(CASE WHEN events_table.level = 'Международный' AND data_table.result = 'Сертификат участника' AND  strftime('%m', events_table.result_date) = '03'  THEN 1 ELSE 0 END) AS c153,
                        SUM(CASE WHEN events_table.level = 'Международный' AND data_table.result != 'Сертификат участника' AND  strftime('%m', events_table.result_date) = '03'  THEN 1 ELSE 0 END) AS c163,
                        SUM(CASE WHEN data_table.result = 'Сертификат участника'  THEN 1 ELSE 0 END) AS i1,
                        SUM(CASE WHEN data_table.result != 'Сертификат участника'  THEN 1 ELSE 0 END) AS i2
                    FROM
                        data_table
                    JOIN spisok ON data_table.id_spisok = spisok.id
                    JOIN spisok_in_studio ON data_table.id_spisok_in_studio = spisok_in_studio.id
                    JOIN teacher ON spisok_in_studio.pedagog = teacher.id
                    JOIN events_table ON data_table.id_events_table = events_table.id
                    WHERE  (events_table.result_date BETWEEN ? AND ?)  and  data_table.result != ' '
                    '''
        if user != 'admin':
            text += ''' 
           and teacher.FIO = ? 
           '''

        text += '''
                    GROUP BY
                        spisok_in_studio.napravlenie, teacher_name
                    ORDER BY
                        spisok_in_studio.napravlenie, teacher_name;

                '''

        if user != 'admin':
            cursor_db.execute(text, ('2025-01-01', '2025-03-31', user))
        else:
            cursor_db.execute(text, ('2025-01-01', '2025-03-31'))

        # Извлечение данных
        data = cursor_db.fetchall()
        columns = [description[0] for description in cursor_db.description]

    except sqlite3.Error as e:
         flash(f'Ошибка при работе с базой данных: {e}', 'error')
    #
    finally:
        # Закрытие соединения с базой данных
        cursor_db.close()
        db_lp.close()

        # Открываем существующий файл Excel
        workbook = openpyxl.load_workbook('templates\Otchet_1.xlsx')
        sheet = workbook.active

        napr ={ n[0] for n in data}
        # Заполнение данных
        if kvartal==1:
            mes= ['Январь', 'Февраль', 'Март']
        for i in range(8):
            for j in range(3):
                sheet.cell(row=4, column=3+(j*2+i*6), value=mes[j])

        row_start = 6  # Начинаем с третьей строки
        for n in napr:

            npp = 1
            sheet.cell(row=row_start, column=1, value=n)
            row_start += 1
            for row in data:
                # Предположим, что values это список из двух чисел
                sheet.cell(row=row_start, column=1, value=npp)
                sheet.cell(row=row_start, column=2, value=row[1])  # Например, имя педагога
                sheet.cell(row=row_start, column=3, value=row[2])  # 1
                sheet.cell(row=row_start, column=4, value=row[3])  #  1
                sheet.cell(row=row_start, column=5, value=row[4])  #  2
                sheet.cell(row=row_start, column=6, value=row[5])  # 2
                sheet.cell(row=row_start, column=7, value=row[6])  # 3
                sheet.cell(row=row_start, column=8, value=row[7])  # 3
                sheet.cell(row=row_start, column=9, value=row[8])  # 3
                sheet.cell(row=row_start, column=10, value=row[9])  # 3
                sheet.cell(row=row_start, column=11, value=row[10])  # 3
                sheet.cell(row=row_start, column=12, value=row[11])  # 3
                sheet.cell(row=row_start, column=13, value=row[12])  # 3
                sheet.cell(row=row_start, column=14, value=row[13])  # 3
                sheet.cell(row=row_start, column=15, value=row[14])  # 3
                sheet.cell(row=row_start, column=16, value=row[15])  # 3
                sheet.cell(row=row_start, column=17, value=row[16])  # 3
                sheet.cell(row=row_start, column=18, value=row[17])  # 3
                sheet.cell(row=row_start, column=19, value=row[18])  # 3
                sheet.cell(row=row_start, column=20, value=row[19])  # 3
                sheet.cell(row=row_start, column=21, value=row[20])  # 3
                sheet.cell(row=row_start, column=22, value=row[21])  # 3
                sheet.cell(row=row_start, column=23, value=row[22])  # 3
                sheet.cell(row=row_start, column=24, value=row[23])  # 3
                sheet.cell(row=row_start, column=25, value=row[24])  # 3
                sheet.cell(row=row_start, column=26, value=row[25])  # 3
                sheet.cell(row=row_start, column=27, value=row[26])  # 3
                sheet.cell(row=row_start, column=28, value=row[27])  # 3
                sheet.cell(row=row_start, column=29, value=row[28])  # 3
                sheet.cell(row=row_start, column=30, value=row[29])  # 3
                sheet.cell(row=row_start, column=31, value=row[30])  # 3
                sheet.cell(row=row_start, column=32, value=row[31])  # 3
                sheet.cell(row=row_start, column=33, value=row[32])  # 3
                sheet.cell(row=row_start, column=34, value=row[33])  # 3
                sheet.cell(row=row_start, column=35, value=row[34])  # 3
                sheet.cell(row=row_start, column=36, value=row[35])  # 3
                sheet.cell(row=row_start, column=37, value=row[36])  # 3
                sheet.cell(row=row_start, column=38, value=row[37])  # 3
                sheet.cell(row=row_start, column=39, value=row[38])  # 3
                sheet.cell(row=row_start, column=40, value=row[39])  # 3
                sheet.cell(row=row_start, column=41, value=row[40])  # 3
                sheet.cell(row=row_start, column=42, value=row[41])  # 3
                sheet.cell(row=row_start, column=43, value=row[42])  # 3
                sheet.cell(row=row_start, column=44, value=row[43])  # 3
                sheet.cell(row=row_start, column=45, value=row[44])  # 3
                sheet.cell(row=row_start, column=46, value=row[45])  # 3
                sheet.cell(row=row_start, column=47, value=row[46])  # 3
                sheet.cell(row=row_start, column=48, value=row[47])  # 3
                sheet.cell(row=row_start, column=49, value=row[48])  # 3
                sheet.cell(row=row_start, column=50, value=row[49])  # 3
                sheet.cell(row=row_start, column=51, value=row[50])  # 3
                sheet.cell(row=row_start, column=52, value=row[51])  # 3
                sheet.cell(row=row_start, column=53, value=row[51]+row[50])  # 3
                row_start += 1
                npp += 1

        # Сохраняем изменения в файл
        workbook.save(file_path)
        workbook.close()

        return file_path

if __name__ == '__main__':
    update_excel_template('admin')


