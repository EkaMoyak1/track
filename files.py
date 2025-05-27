import openpyxl
import datetime
import os
import sqlite3
from flask import flash

def update_excel_template(user, polugodie, year_1, year_2):
    # Формирование имени файла
    file_name = f'v_{datetime.datetime.today().strftime("%Y-%m-%d_%H-%M-%S")}.xlsx'
    # Получение текущей директории файла и добавление подкаталога `tmp`
    current_dir = os.path.abspath(os.path.dirname(__file__))
    tmp_dir = os.path.join(current_dir, 'tmp')
    # Убедитесь, что директория `tmp` существует
    os.makedirs(tmp_dir, exist_ok=True)
    # Полный путь к файлу в подкаталоге `tmp`
    file_path = os.path.join(tmp_dir, file_name)

    try:
        # Подключение к базе данных
        db_lp = sqlite3.connect('date_source.db')
        cursor_db = db_lp.cursor()

        # SQL-запрос с динамическим условием для пользователя
        base_query = '''
            SELECT
                spisok_in_studio.napravlenie,
                teacher.FIO AS teacher_name,
                events_table.level,
                strftime('%m', events_table.result_date) AS month,
                data_table.result,
                COUNT(*) AS count
            FROM
                data_table
            JOIN spisok ON data_table.id_spisok = spisok.id
            JOIN spisok_in_studio ON data_table.id_spisok_in_studio = spisok_in_studio.id
            JOIN teacher ON spisok_in_studio.pedagog = teacher.id
            JOIN events_table ON data_table.id_events_table = events_table.id
            WHERE
                events_table.result_date BETWEEN ? AND ?
                AND data_table.result != ' '
        '''
        if user != 'admin':
            base_query += ' AND teacher.FIO = ?'

        base_query += '''
            GROUP BY
                spisok_in_studio.napravlenie,
                teacher.FIO,
                events_table.level,
                strftime('%m', events_table.result_date),
                data_table.result
            ORDER BY
                spisok_in_studio.napravlenie,
                teacher.FIO,
                events_table.level,
                strftime('%m', events_table.result_date);
        '''

        # Параметры для запроса
        if polugodie == 2:
            dbeg = f'{year_2}-01-01'
            dfin = f'{year_2}-05-31'
        elif polugodie == 1:
            dbeg = f'{year_1}-09-01'
            dfin = f'{year_1}-12-31'
        else:
            dbeg = f'{year_1}-09-01'
            dfin = f'{year_2}-05-31'
        params = (dbeg, dfin)
        if user != 'admin':
            params += (user,)

        # Выполнение запроса
        cursor_db.execute(base_query, params)
        data = cursor_db.fetchall()

        # Обработка данных
        result_data = {}
        for row in data:

            napravlenie, teacher_name, level, month, result, count = row
            if napravlenie not in result_data:
                result_data[napravlenie] = {}
            if teacher_name not in result_data[napravlenie]:
                result_data[napravlenie][teacher_name] = {}
            if level not in result_data[napravlenie][teacher_name]:
                result_data[napravlenie][teacher_name][level] = {}
            if month not in result_data[napravlenie][teacher_name][level]:
                result_data[napravlenie][teacher_name][level][month] = {'с': 0, 'р': 0}
            result_data[napravlenie][teacher_name][level][month]['с' if result == 'Сертификат участника' else 'р'] += count


        if polugodie == 2:
            # Открываем существующий файл Excel
            workbook = openpyxl.load_workbook('templates/Otchet_2_polugodie.xlsx')
            months = ['01', '02', '03', '04', '05']
        elif polugodie == 1:
            # Открываем существующий файл Excel
            workbook = openpyxl.load_workbook('templates/Otchet_1_polugodie.xlsx')
            months = ['09', '10', '11', '12']

        sheet = workbook.active
        print(result_data)

        # Заполняем данные
        row_start = 6  # Начинаем с 6 строки

        levels = ['Центровский', 'Городской', 'Районный', 'Республиканский', 'Региональный', 'Межрегиональный', 'Всероссийский', 'Международный']

        # # Заголовки месяцев
        # for i, level in enumerate(levels):
        #     for j, month in enumerate(months):
        #         sheet.cell(row=4, column=3 + (j * 2 + i * 6), value=month)
        total_c = 0
        total_p = 0
        # Заполняем данные
        for napravlenie in sorted(result_data.keys()):
            npp = 1
            # sheet.cell(row=row_start, column=1, value=napravlenie)
            # row_start += 1
            for teacher_name in sorted(result_data[napravlenie].keys()):
                sheet.cell(row=row_start, column=1, value=npp)
                sheet.cell(row=row_start, column=2, value=teacher_name)
                col_index = 3
                for i, level in enumerate(levels):
                    for j, month in enumerate(months):
                        # col_index = 3 + (j * 2 + i*len(months)*2)
                        # сертификаты
                        res_c = result_data[napravlenie][teacher_name].get(level, {}).get(month, {}).get('с', 0)
                        res_p = result_data[napravlenie][teacher_name].get(level, {}).get(month, {}).get('р', 0)
                        sheet.cell(row=row_start, column=col_index, value=res_c)
                        sheet.cell(row=row_start, column=col_index + 1, value=res_p)
                        total_c += res_c
                        total_p += res_p
                        col_index += 2
                sheet.cell(row=row_start, column=col_index, value=total_c)
                sheet.cell(row=row_start, column=col_index + 1, value=total_p)
                sheet.cell(row=row_start, column=col_index + 2, value=total_c+total_p)
                row_start += 1
                npp += 1

        # Сохраняем изменения в файл
        workbook.save(file_path)
        workbook.close()
        return file_path

    except sqlite3.Error as e:
        flash(f'Ошибка при работе с базой данных: {e}', 'error')
        return None

    finally:
        # Закрытие соединения с базой данных
        cursor_db.close()
        db_lp.close()

if __name__ == '__main__':
    update_excel_template('admin')